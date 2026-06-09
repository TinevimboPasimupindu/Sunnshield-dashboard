from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from pydantic import BaseModel
from typing import Optional
import models
from datetime import datetime, timedelta
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
import io
import openpyxl

router = APIRouter(prefix="/api/batches", tags=["batches"])

class BatchCreate(BaseModel):
    notes: Optional[str] = None

@router.get("/")
def get_batches(db: Session = Depends(get_db)):
    batches = db.query(models.Batch).order_by(models.Batch.created_at.desc()).all()
    result = []
    for batch in batches:
        total_orders = len(batch.batch_orders)
        approved = sum(1 for bo in batch.batch_orders if bo.status == models.BatchOrderStatusEnum.APPROVED)
        rejected = sum(1 for bo in batch.batch_orders if bo.status == models.BatchOrderStatusEnum.REJECTED)
        total_amount_usd = sum(
            bo.order.amount for bo in batch.batch_orders
            if bo.order.currency == models.CurrencyEnum.USD
        )
        total_amount_zwl = sum(
            bo.order.amount for bo in batch.batch_orders
            if bo.order.currency == models.CurrencyEnum.ZWL
        )
        result.append({
            "id": batch.id,
            "batch_number": batch.batch_number,
            "status": batch.status,
            "notes": batch.notes,
            "submitted_at": batch.submitted_at,
            "created_at": batch.created_at,
            "total_orders": total_orders,
            "approved": approved,
            "rejected": rejected,
            "total_amount_usd": round(total_amount_usd, 2),
            "total_amount_zwl": round(total_amount_zwl, 2),
        })
    return result

@router.post("/")
def create_batch(data: BatchCreate, db: Session = Depends(get_db)):
    count = db.query(models.Batch).count()
    batch_number = f"BATCH-{datetime.now().year}-{str(count + 1).zfill(3)}"
    batch = models.Batch(
        batch_number=batch_number,
        notes=data.notes,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch

@router.get("/open")
def get_open_batch(db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(
        models.Batch.status == models.BatchStatusEnum.OPEN
    ).first()
    if not batch:
        return {"batch": None}
    return {
        "batch": {
            "id": batch.id,
            "batch_number": batch.batch_number,
            "status": batch.status,
            "total_orders": len(batch.batch_orders),
        }
    }

@router.post("/import-global")
async def import_global(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    try:
        ec_col = headers.index("Ec number")
        status_col = headers.index("Status")
        message_col = headers.index("Message") if "Message" in headers else None
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid file format — missing required columns")

    approved_count = 0
    rejected_count = 0
    not_found = []
    updated_batches = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        ec_number = str(row[ec_col]).strip() if row[ec_col] else None
        status = str(row[status_col]).strip() if row[status_col] else None
        message = str(row[message_col]).strip() if message_col is not None and row[message_col] else None

        if not ec_number:
            continue

        batch_order = db.query(models.BatchOrder).join(models.Order).join(models.Client).filter(
            models.Client.ec_number == ec_number,
            models.BatchOrder.status == models.BatchOrderStatusEnum.PENDING
        ).first()

        if not batch_order:
            not_found.append(ec_number)
            continue

        if status == "SUCCESS":
            batch_order.status = models.BatchOrderStatusEnum.APPROVED
            approved_count += 1
        elif status == "FAILED":
            batch_order.status = models.BatchOrderStatusEnum.REJECTED
            batch_order.rejection_reason = message
            rejected_count += 1

        updated_batches.add(batch_order.batch_id)

    for batch_id in updated_batches:
        batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
        if not batch:
            continue
        all_statuses = [bo.status for bo in batch.batch_orders]
        pending = sum(1 for s in all_statuses if s == models.BatchOrderStatusEnum.PENDING)
        approved = sum(1 for s in all_statuses if s == models.BatchOrderStatusEnum.APPROVED)
        rejected = sum(1 for s in all_statuses if s == models.BatchOrderStatusEnum.REJECTED)

        if pending > 0 or rejected > 0:
            batch.status = models.BatchStatusEnum.IN_PROGRESS
        elif approved > 0 and rejected == 0 and pending == 0:
            batch.status = models.BatchStatusEnum.APPROVED

    db.commit()

    return {
        "message": "Global import processed",
        "approved": approved_count,
        "rejected": rejected_count,
        "batches_updated": len(updated_batches),
        "not_found": not_found,
    }

@router.post("/{batch_id}/add-order/{order_id}")
def add_order_to_batch(batch_id: int, order_id: int, db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    order = db.query(models.Order).filter(models.Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    batch_order = models.BatchOrder(
        batch_id=batch_id,
        order_id=order_id,
    )
    db.add(batch_order)
    db.commit()
    return {"message": "Order added to batch"}

@router.put("/{batch_id}/status")
def update_batch_status(batch_id: int, status: str, db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    try:
        batch.status = models.BatchStatusEnum(status)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid status")
    if status == "SUBMITTED":
        batch.submitted_at = datetime.utcnow()
    db.commit()
    return {"message": "Status updated"}

@router.get("/{batch_id}/orders")
def get_batch_orders(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    orders = []
    for bo in batch.batch_orders:
        orders.append({
            "batch_order_id": bo.id,
            "order_id": bo.order.id,
            "status": bo.status,
            "rejection_reason": bo.rejection_reason,
            "adjusted_term_months": bo.adjusted_term_months,
            "adjusted_instalment": bo.adjusted_instalment,
            "client": {
                "full_name": bo.order.client.full_name,
                "ec_number": bo.order.client.ec_number,
                "id_number": bo.order.client.id_number,
            },
            "amount": bo.order.amount,
            "currency": bo.order.currency,
            "term_months": bo.order.term_months,
            "monthly_instalment": bo.order.monthly_instalment,
            "reference_number": bo.order.reference_number,
        })
    return {"batch_number": batch.batch_number, "status": batch.status, "orders": orders}

@router.get("/{batch_id}/export")
def export_batch(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["Reference", "IdNumber", "EcNumber", "Type", "StartDate", "EndDate", "Amount"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    col_widths = [15, 20, 15, 10, 15, 15, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    for bo in batch.batch_orders:
        order = bo.order
        client = order.client
        start_date = order.start_date or (order.created_at.strftime("%d/%b/%Y") if order.created_at else datetime.utcnow().strftime("%d/%b/%Y"))
        end_date = order.end_date or ""
        ws.append([
            order.reference_number or "",
            client.id_number,
            client.ec_number,
            "NEW",
            start_date,
            end_date,
            order.monthly_instalment,
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{batch.batch_number}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/{batch_id}/import-response")
async def import_response(batch_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    try:
        ec_col = headers.index("Ec number")
        status_col = headers.index("Status")
        message_col = headers.index("Message") if "Message" in headers else None
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid file format")

    approved_count = 0
    rejected_count = 0
    not_found = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        ec_number = str(row[ec_col]).strip() if row[ec_col] else None
        status = str(row[status_col]).strip() if row[status_col] else None
        message = str(row[message_col]).strip() if message_col is not None and row[message_col] else None

        if not ec_number:
            continue

        batch_order = db.query(models.BatchOrder).join(models.Order).join(models.Client).filter(
            models.BatchOrder.batch_id == batch_id,
            models.Client.ec_number == ec_number
        ).first()

        if not batch_order:
            not_found.append(ec_number)
            continue

        if status == "SUCCESS":
            batch_order.status = models.BatchOrderStatusEnum.APPROVED
            approved_count += 1
        elif status == "FAILED":
            batch_order.status = models.BatchOrderStatusEnum.REJECTED
            batch_order.rejection_reason = message
            rejected_count += 1

    all_statuses = [bo.status for bo in batch.batch_orders]
    pending = sum(1 for s in all_statuses if s == models.BatchOrderStatusEnum.PENDING)
    approved = sum(1 for s in all_statuses if s == models.BatchOrderStatusEnum.APPROVED)
    rejected = sum(1 for s in all_statuses if s == models.BatchOrderStatusEnum.REJECTED)

    if pending > 0 or rejected > 0:
        batch.status = models.BatchStatusEnum.IN_PROGRESS
    elif approved > 0 and rejected == 0 and pending == 0:
        batch.status = models.BatchStatusEnum.APPROVED

    db.commit()

    return {
        "message": "Response file processed",
        "approved": approved_count,
        "rejected": rejected_count,
        "not_found": not_found,
    }

@router.delete("/{batch_id}")
def delete_batch(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    # Block deletion if batch has approved orders
    approved = sum(1 for bo in batch.batch_orders 
                   if bo.status == models.BatchOrderStatusEnum.APPROVED)
    if approved > 0:
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete a batch with approved orders"
        )
    
    db.query(models.BatchOrder).filter(
        models.BatchOrder.batch_id == batch_id
    ).delete()
    db.delete(batch)
    db.commit()
    return {"message": "Batch deleted"}